from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import StreamingResponse
import csv
import io
import re
from decimal import Decimal, InvalidOperation
import requests
from openpyxl import Workbook
from openpyxl.styles import Font

app = FastAPI()

# =============================
# Настройки
# =============================
YANDEX_API_KEY = "os.getenv("YANDEX_API_KEY")"
FOLDER_ID = "b1gj8gdme1f9bbsk0vgh"
YANDEX_URL = "https://llm.api.cloud.yandex.net/foundationModels/v1/completion"

CSV_HEADERS = [
    "Наименование",
    "Ед. изм",
    "Кол-во",
    "Цена за ед., руб",
    "Сумма, руб",
]

# =============================
# Запрос к YandexGPT
# =============================
def ask_yandex_gpt(user_text: str) -> str:
    headers = {
        "Authorization": f"Api-Key {YANDEX_API_KEY}",
        "Content-Type": "application/json",
    }
    data = {
        "modelUri": f"gpt://{FOLDER_ID}/yandexgpt/latest",
        "completionOptions": {
            "stream": False,
            "temperature": 0.3,
            "maxTokens": "4000",
        },
        "messages": [
            {
                "role": "system",
                "text": (
                    "Ты — профессиональный инженер-сметчик по ремонту квартир. "
                    "Составь смету по тексту пользователя. "
                    "Верни только CSV без пояснений, без Markdown и без кавычек. "
                    "Первая строка: Наименование;Ед. изм;Кол-во;Цена за ед., руб;Сумма, руб. "
                    "Разделитель — точка с запятой."
                ),
            },
            {
                "role": "user",
                "text": user_text,
            },
        ],
    }
    response = requests.post(YANDEX_URL, headers=headers, json=data, timeout=120)
    response.raise_for_status()
    result = response.json()
    try:
        return result["result"]["alternatives"][0]["message"]["text"]
    except (KeyError, IndexError, TypeError) as e:
        raise ValueError("YandexGPT вернул неожиданный формат") from e

# =============================
# Числа
# =============================
def parse_number(value: str) -> Decimal:
    value = value.strip().replace(" ", "").replace("₽", "")
    if "," in value and "." in value:
        if value.rfind(",") > value.rfind("."):
            value = value.replace(".", "").replace(",", ".")
        else:
            value = value.replace(",", "")
    else:
        value = value.replace(",", ".")
    try:
        return Decimal(value)
    except InvalidOperation as e:
        raise ValueError(f"Некорректное число: {value}") from e

# =============================
# Excel
# =============================
def create_excel(csv_text: str) -> bytes:
    csv_text = csv_text.strip()
    csv_text = re.sub(r"^```(?:csv)?\s*", "", csv_text, flags=re.IGNORECASE)
    csv_text = re.sub(r"\s*```$", "", csv_text)

    reader = csv.reader(io.StringIO(csv_text), delimiter=";")
    rows = list(reader)
    if not rows:
        raise ValueError("Пустой CSV")

    first_row = [cell.strip() for cell in rows[0]]
    if first_row != CSV_HEADERS:
        raise ValueError("Неверные заголовки CSV")

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.append(CSV_HEADERS)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    total = Decimal("0")

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or not any(cell.strip() for cell in row):
            continue
        if len(row) != 5:
            raise ValueError(f"Строка {row_number} должна содержать 5 столбцов")

        name = row[0].strip()
        unit = row[1].strip()
        quantity = parse_number(row[2])
        price = parse_number(row[3])
        amount = parse_number(row[4])

        ws.append([name, unit, float(quantity), float(price), float(amount)])
        total += amount

    if ws.max_row == 1:
        raise ValueError("Нет позиций")

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.cell(row=total_row, column=5, value=float(total))

    for cell in ws[total_row]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# =============================
# API
# =============================
@app.post("/generate")
async def generate(request: Request):
    body = await request.json()
    user_text = body.get("text", "")

    if not user_text or not user_text.strip():
        raise HTTPException(status_code=400, detail="Пустой текст")

    try:
        csv_text = ask_yandex_gpt(user_text.strip())
        excel_bytes = create_excel(csv_text)

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=smeta.xlsx"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/health")
async def health():
    return {"status": "ok"}